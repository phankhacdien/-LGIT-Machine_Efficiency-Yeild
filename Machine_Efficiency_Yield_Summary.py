# -*- coding: utf-8 -*-
"""
Created on Wed Oct 06 09:59:38 2021

@author: Sherlock
"""

from tkinter import *
from tkinter.filedialog import *
import pandas as pd
from statistics import *
from datetime import *
from datetime import datetime, timedelta
import time
import numpy as np
import xlsxwriter
from openpyxl import *
# import os

def TimeIs(func):
    def wrap(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = time.time()
        print("[ ", func.__name__, " process in ", end-start, " ]")
        return result
    return wrap


def SelectExcelFile(idex_title):
    titles = {1: "Take Lot_time data file", 2: "Take HM_ERROR data file"}
    excel_filetypes = [('All files', '*'), ('CSV files', '*.CSV*'), ('Excel files', '*.xlsx*')]
    Tk().withdraw()
    file_path = askopenfilename(initialdir='./', title=titles[idex_title], filetypes=excel_filetypes)
    return file_path


def InsertRow(row_number, df, row_value):
    start_upper = 0         # Starting value of upper half 
    end_upper = row_number  # End value of upper half 
    
    start_lower = row_number    # Start value of lower half 
    end_lower = df.shape[0]     # End value of lower half 
    
    upper_half = [*range(start_upper, end_upper, 1)]    # Create a list of upper_half index 
    lower_half = [*range(start_lower, end_lower, 1)]    # Create a list of lower_half index 
    
    lower_half = [x.__add__(1) for x in lower_half]     # Increment the value of lower half by 1 
    index_ = upper_half + lower_half                    # Combine the two lists 
    df.index = index_                                   # Update the index of the dataframe 

    df.loc[row_number] = row_value  # Insert a row at the end 
    df = df.sort_index()            # Sort the index labels 
    return df


def FindTestTime(data):
    test_time = list()
    header_site = [i for i in data.columns if "Site" in i]
    header_site = header_site[:8]
    for header in header_site:
        tt_med = median(data[header])
        test_time.append(tt_med)
    return test_time
    

def TakeRetestCountTable(LT, list_items_fail):
    retest_count = pd.DataFrame(columns=['initial', '1st_retest', '2nd_retest', '3rd_retest'], index=list_items_fail)
    
    total_test_count = LT.shape
    total_test_count = total_test_count[0] + 1
    
    sort_data = LT.sort_values(by=['Barcode','Time'],ascending=[True,True])
    rtc_data = sort_data.reset_index(drop=True)
    rtc_data = rtc_data[['Barcode', 'ERROR']]
    rtc_data['one'] = 1
    rtc_data['retest_count'] = rtc_data.groupby(by=['Barcode'])['one'].cumsum()
    
    for i in range(1,5):
        rtc = rtc_data[rtc_data['retest_count'] == i]
        if i == 1:
            test_turn = "initial"
        elif i == 2:
            test_turn = "1st_retest"
        elif i == 3:
            test_turn = "2nd_retest"
        else:
            test_turn = "3rd_retest"
        for j in range(0,len(list_items_fail)):
            rtc_ = rtc[rtc['ERROR'] == list_items_fail[j]]
            # idx = retest_count[retest_count['fail_items'] == list_items_fail[j]].index
            retest_count.loc[list_items_fail[j],[test_turn]] = len(rtc_['Barcode'])
            
    for item in list_items_fail:
        if ("DP" in item) or ("LCB" in item) or ("PASS" in item):
            retest_count.loc[[item],['1st_retest', '2nd_retest', '3rd_retest']] = 0  
    
    retest_count["1st_retest_PASS"] = retest_count['initial'] - retest_count['1st_retest']
    retest_count["2nd_retest_PASS"] = retest_count['1st_retest'] - retest_count['2nd_retest']
    retest_count["3rd_retest_PASS"] = retest_count['2nd_retest'] - retest_count['3rd_retest']
    
    retest_count[retest_count[['1st_retest_PASS', '2nd_retest_PASS', '3rd_retest_PASS']]<0] = 0
    
    retest_count["Retest Count Rate(%)"] = (retest_count['initial']+retest_count['1st_retest']+retest_count['2nd_retest'])/total_test_count*100
    retest_count["Retest Pass(%)"] = (retest_count['1st_retest_PASS']+retest_count['2nd_retest_PASS']+retest_count['3rd_retest_PASS'])/total_test_count*100
    retest_count = retest_count.sort_values(by=['initial'], ascending=False)
    retest_count = retest_count.reset_index()
    
    lcb_dp_cnt = retest_count[retest_count['index'].str.contains('LCB|DP')]
    retest_count = retest_count[~retest_count['index'].str.contains('LCB|DP')]
    
    retest_count = retest_count.reset_index(drop=True)
    lcb_dp_cnt = lcb_dp_cnt.reset_index(drop=True)
    return retest_count, lcb_dp_cnt, total_test_count
    

def ConvertTime(inputdata):
    data = inputdata
    data['Convert time'] = ""
    data['Check start/stop'] = ""
    # data.insert(0, 'Convert time',"")
    # data.insert(0, 'Check start/stop',"")
    
    for i in range(0, data.shape[0]):
        try:
            T = data['Time'][i]
        except:
            T = data['time'][i]
            
        try:
            var = datetime.strptime(T[1:T.index(".")], "%Y-%m-%d %H:%M:%S")
        except:
            var = datetime.strptime(T[1:T.rfind(":")], "%Y-%m-%d %H:%M:%S")
            
        data.loc[i,['Convert time']] = var
    return data


def MachineEfficiency(Lot_time, Human_error):
    eff_LT = ConvertTime(Lot_time)
    eff_HM = ConvertTime(Human_error)
    
    eff_LT = eff_LT.sort_values(by=['Convert time'], ascending=True)
    eff_LT = eff_LT.reset_index(drop=True)
    eff_LT = eff_LT[['Check start/stop','Convert time', 'LotNum']]
    
    eff_LT["LotNum+1"] = eff_LT['LotNum'].shift(periods=1)
    eff_LT["LotNum-1"] = eff_LT['LotNum'].shift(periods=-1)
    eff_LT['Check start/stop'].loc[(eff_LT['LotNum'] != eff_LT['LotNum+1']) & (eff_LT['LotNum'] == eff_LT['LotNum-1'])] = "LOT start"
    eff_LT['Check start/stop'].loc[(eff_LT['LotNum'] == eff_LT['LotNum+1']) & (eff_LT['LotNum'] != eff_LT['LotNum-1'])] = "LOT stop"
    eff_LT = eff_LT[['Check start/stop','Convert time', 'LotNum']]
    
    # for j in range(0,len(eff_LT)-1):
    #     if j == 0:
    #         eff_LT['Check start/stop'][j] = 'LOT start'
    #     elif ((eff_LT['LotNum'][j] == eff_LT['LotNum'][j+1]) and (eff_LT['LotNum'][j] != eff_LT['LotNum'][j-1])):
    #         eff_LT['Check start/stop'][j] = 'LOT start'
    #     elif ((eff_LT['LotNum'][j] == eff_LT['LotNum'][j-1]) and (eff_LT['LotNum'][j] != eff_LT['LotNum'][j+1])):
    #         eff_LT['Check start/stop'][j] = 'LOT stop'
    #     elif j+1 == len(eff_LT)-1:
    #         eff_LT['Check start/stop'][j+1] = 'LOT stop'
    #     else:
    #         eff_LT['Check start/stop'][j] = ''
    
    eff_LT = eff_LT[eff_LT['Check start/stop'] != ""]
    eff_LT = eff_LT.reset_index(drop=True)
    eff_LT["Run_time"] = eff_LT['Convert time'].diff()
    
    for i in range(0,len(eff_LT)):
        if i % 2 == 0:
            eff_LT.loc[i, ['Run_time']] = np.nan
    
    sum_runtime = eff_LT['Run_time'].sum()
    
    eff_all = eff_LT[["Convert time", "Check start/stop"]]
    eff_all.insert(1, "Duration_time", "")
    eff_all = eff_all.rename(columns= {"Check start/stop": "event"})
    
    eff_HM = eff_HM[['Convert time', 'event', 'Action Time']]
    eff_HM.insert(1, "Duration_time", "")
    eff_HM['Duration_time'] = pd.to_timedelta(arg=eff_HM['Action Time'].astype('float'), unit='sec')
    eff_HM = eff_HM[['Convert time', 'Duration_time', 'event']]
    
    eff_all = pd.concat([eff_all,eff_HM])
    eff_all = eff_all.sort_values(by=['Convert time'], ascending=True)
    eff_all = eff_all.reset_index(drop=True)
    
    eff_all['filter'] = np.nan
    eff_all['filter'].loc[eff_all['event'] == "LOT start"] = 1
    eff_all['filter'].loc[eff_all['event'] == "LOT stop"] = 2
    eff_all['filter'] = eff_all[['filter']].fillna(method='ffill')
    eff_all['filter'].loc[eff_all['event'] == "LOT stop"] = 1
    eff_all['filter'].loc[eff_all['filter'] == 2] = np.nan
    eff_all = eff_all.dropna(axis=0, how='any')
    eff_all = eff_all.reset_index(drop=True)
    return eff_all, sum_runtime
    

def AppendDfToExcel(filename, df, sheet_name='Sheet1', startrow=None, startcol=None, 
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    # try:
    #     FileNotFoundError
    # except NameError:
    #     FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
        
    if startcol is None:
        startcol = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, header=False, index=False, startrow=startrow, startcol=startcol, **to_excel_kwargs)

    # save the workbook
    writer.save()
    

def YieldSummary(data):
    yield_dt = data.sort_values(by=['Time'], ascending=False)
    yield_dt = yield_dt.drop_duplicates()
    yield_dt = yield_dt.reset_index(drop=True)
    input_qty = len(yield_dt)
    fail_items = [i for i in yield_dt['ERROR'].unique() if i != '-']
    summary_dt = pd.DataFrame(index=fail_items, columns=["Quantity"])
    for item in fail_items:
        yield_dt_ = yield_dt[yield_dt['ERROR'] == item]
        summary_dt.loc[item,'Quantity'] = len(yield_dt_)
    summary_dt = summary_dt.sort_values(by=['Quantity'], ascending=False)
    summary_dt = summary_dt.reset_index()
    summary_dt["Rate(%)"] = summary_dt['Quantity']/input_qty
    summary_dt = summary_dt.rename(columns={"index":"fail_items"})
    summary_dt['Rate(%)'] = summary_dt['Rate(%)'].astype(float).map(lambda n :'{:.2%}'.format(n))
    return summary_dt
    

def ComputeUPH(data):
    tact_time_med = median(data['Tact_Time(Sec)'])   
    test_time_max = max(FindTestTime(data))
    # handling_time = float(tact_time_med)-float(test_time_longest)
    # pure_UPH = 3600/float(tact_time_med) # item did not use yet
    return test_time_max, float(tact_time_med)-float(test_time_max)
            

def ReadAndPreProcessData(path):
    if path.endswith('.CSV') or path.endswith('.csv'):
        data = pd.read_csv(path, encoding='cp949')
    else:
        data = pd.read_excel(path)
    
    if "LOT_TIME" in path:
        data = data[data['Time'] != "Time"]
        data = data.reset_index(drop=True)
    else:
        data = data[data['Event'] == "ERROR"]
        data = data.reset_index(drop=True)
    return data
    

@TimeIs
def YieldSummaryMain():
    # Read log file
    path_LT = SelectExcelFile(1)
    data_LT = ReadAndPreProcessData(path_LT)
    summary_tbl = YieldSummary(data_LT)
    
    # Exporting result
    print("[ >>> Exporting result >>> ]")
    summary_tbl.to_excel("Yield Summary.xlsx", sheet_name='result', na_rep='', index=False, startrow=0, startcol=1)
    

@TimeIs
def MachineEfficiencyMain():
    # Read log file
    path_LT = SelectExcelFile(1)
    path_HM = SelectExcelFile(2)
    
    data_LT = ReadAndPreProcessData(path_LT)
    data_HM = ReadAndPreProcessData(path_HM)
    
    # compute UPH
    [test_time_longest, handling_time] = ComputeUPH(data_LT)
    
    # Retest rate
    fail_items = [i for i in data_LT['ERROR'].unique() if i != '-']
    
    [retest_count_tbl, lcb_dp_tbl, total_test_count] = TakeRetestCountTable(data_LT, fail_items)
    
    # M/C Eff
    [eff_tbl, total_run_time] = MachineEfficiency(data_LT, data_HM)
    
    # Exporting result
    print("[ >>> Exporting result >>> ]")
    wb = load_workbook("./template.xlsx")
    ws = wb['result']
    ws['C7'] = test_time_longest
    ws['D7'] = handling_time
    ws['D19'] = total_test_count
    ws['D92'] = total_run_time
    ws['C92'] = eff_tbl.loc[0,'Convert time']
    ws['C93'] = eff_tbl.loc[len(eff_tbl)-1,'Convert time']
    wb.save("result.xlsx")
    
    AppendDfToExcel("result.xlsx", retest_count_tbl['index'], sheet_name="result", startrow=23, startcol=1)
    AppendDfToExcel("result.xlsx", retest_count_tbl['initial'], sheet_name="result", startrow=23, startcol=7)
    AppendDfToExcel("result.xlsx", retest_count_tbl['1st_retest'], sheet_name="result", startrow=23, startcol=9)
    AppendDfToExcel("result.xlsx", retest_count_tbl['2nd_retest'], sheet_name="result", startrow=23, startcol=11)
    AppendDfToExcel("result.xlsx", retest_count_tbl['3rd_retest'], sheet_name="result", startrow=23, startcol=13)
    AppendDfToExcel("result.xlsx", lcb_dp_tbl['index'], sheet_name="result", startrow=75, startcol=1)
    AppendDfToExcel("result.xlsx", lcb_dp_tbl['initial'], sheet_name="result", startrow=75, startcol=7)
    AppendDfToExcel("result.xlsx", eff_tbl['Duration_time'], sheet_name="result", startrow=96, startcol=3)
    AppendDfToExcel("result.xlsx", eff_tbl['Convert time'], sheet_name="result", startrow=96, startcol=2)
    AppendDfToExcel("result.xlsx", eff_tbl['event'], sheet_name="result", startrow=96, startcol=4)
    

if __name__ == '__main__':
    while  True:
        run = input("Select running Type ( Yield Summary:1 / Machine efficiency:2 ) : ")
        if (run == "1") or (run == "2"):
            break
        else:
            print("[ Oops! That was no valid running type. Please, try again... ]") 
            
    if run == "1":
        print("[ Yield  Summary start --> ]")
        YieldSummaryMain()
    if run == "2":
        MachineEfficiencyMain()
    
    print("[ DONE !!! (~.~!) ]")
    